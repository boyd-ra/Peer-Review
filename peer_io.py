from __future__ import annotations
from datetime import datetime
import json
import logging
from pathlib import Path
import re
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

import numpy as np
import pydicom
from openpyxl import load_workbook

from peer_helpers import (
    get_ct_row_col_normal,
    get_ipp,
    get_iop,
    normalize_structure_name,
    safe_get,
    sample_dose_to_ct_slice,
)
from peer_models import (
    CTVolume,
    DoseVolume,
    PatientFileDiscovery,
    RTPlanPhase,
    RTStructData,
    StructureGoal,
    StructureSliceContours,
)

try:
    from pydicom.pixels import apply_rescale
except ImportError:
    from pydicom.pixel_data_handlers.util import apply_modality_lut as apply_rescale

try:
    from scipy.ndimage import affine_transform as scipy_affine_transform  # type: ignore
except Exception:  # pragma: no cover - optional runtime dependency
    scipy_affine_transform = None


logger = logging.getLogger(__name__)
_CONSTRAINT_WORKBOOK_SHEETS_CACHE: Dict[Tuple[str, int, int], List[str]] = {}


def _parse_dicom_datetime(date_value: object, time_value: object) -> Optional[datetime]:
    date_text = str(date_value or "").strip()
    if not date_text:
        return None

    time_text = str(time_value or "").strip()
    time_text = time_text.split(".")[0]
    time_text = "".join(ch for ch in time_text if ch.isdigit())
    while len(time_text) < 6:
        time_text += "0"
    time_text = time_text[:6]

    for format_string in ("%Y%m%d%H%M%S", "%Y%m%d"):
        try:
            candidate_text = f"{date_text}{time_text}" if format_string == "%Y%m%d%H%M%S" else date_text
            return datetime.strptime(candidate_text, format_string)
        except ValueError:
            continue
    return None


def _get_dataset_datetime(ds) -> Optional[datetime]:
    for date_field, time_field in (
        ("SeriesDate", "SeriesTime"),
        ("StudyDate", "StudyTime"),
        ("InstanceCreationDate", "InstanceCreationTime"),
        ("ContentDate", "ContentTime"),
    ):
        value = _parse_dicom_datetime(safe_get(ds, date_field, ""), safe_get(ds, time_field, ""))
        if value is not None:
            return value
    return None


def _datetime_sort_key(value: Optional[datetime]) -> Tuple[bool, object]:
    return (value is not None, value or "")


def _extract_referenced_ct_series_uid(ds) -> str:
    for frame_ref in safe_get(ds, "ReferencedFrameOfReferenceSequence", []):
        for study_ref in safe_get(frame_ref, "RTReferencedStudySequence", []):
            for series_ref in safe_get(study_ref, "RTReferencedSeriesSequence", []):
                series_uid = str(safe_get(series_ref, "SeriesInstanceUID", "")).strip()
                if series_uid:
                    return series_uid
    return ""


def _extract_referenced_rtstruct_uid(ds) -> str:
    for item in safe_get(ds, "ReferencedStructureSetSequence", []):
        referenced_uid = str(safe_get(item, "ReferencedSOPInstanceUID", "")).strip()
        if referenced_uid:
            return referenced_uid
    return ""


def _load_registration_transforms(path: str) -> Dict[Tuple[str, str], np.ndarray]:
    try:
        ds = pydicom.dcmread(path, stop_before_pixels=True, force=True)
    except Exception as exc:
        logger.warning("Skipping unreadable REG during folder scan: %s (%s)", path, exc)
        return {}

    frame_to_registered_space: Dict[str, np.ndarray] = {}
    for registration_item in safe_get(ds, "RegistrationSequence", []):
        frame_uid = str(safe_get(registration_item, "FrameOfReferenceUID", "")).strip()
        if not frame_uid:
            continue

        transform_matrix = np.eye(4, dtype=np.float64)
        matrix_found = False
        for matrix_registration_item in safe_get(registration_item, "MatrixRegistrationSequence", []):
            for matrix_item in safe_get(matrix_registration_item, "MatrixSequence", []):
                matrix_payload = safe_get(matrix_item, "FrameOfReferenceTransformationMatrix", None)
                try:
                    transform_matrix = np.asarray(matrix_payload, dtype=np.float64).reshape(4, 4)
                except (TypeError, ValueError):
                    continue
                matrix_found = True
                break
            if matrix_found:
                break

        frame_to_registered_space[frame_uid] = transform_matrix

    pairwise_transforms: Dict[Tuple[str, str], np.ndarray] = {}
    for source_frame_uid, source_to_registered in frame_to_registered_space.items():
        for target_frame_uid, target_to_registered in frame_to_registered_space.items():
            if source_frame_uid == target_frame_uid:
                pairwise_transforms[(source_frame_uid, target_frame_uid)] = np.eye(4, dtype=np.float64)
                continue
            try:
                pairwise_transforms[(source_frame_uid, target_frame_uid)] = (
                    np.linalg.inv(target_to_registered) @ source_to_registered
                )
            except np.linalg.LinAlgError:
                logger.warning(
                    "Skipping singular REG transform in %s between %s and %s",
                    path,
                    source_frame_uid,
                    target_frame_uid,
                )
    return pairwise_transforms


def get_constraints_workbook_path() -> Optional[str]:
    path = Path(__file__).resolve().with_name("constraints.xlsx")
    if path.exists():
        return str(path)
    return None


def _parse_structure_goal_rows(
    fieldnames: List[str],
    rows: List[Dict[str, object]],
) -> Tuple[set[str], dict[str, List[StructureGoal]], List[str]]:
    allowed_names: set[str] = set()
    goals_by_structure: dict[str, List[StructureGoal]] = {}
    structure_order: List[str] = []
    oar_field = next((field for field in fieldnames if normalize_structure_name(field) == "OAR"), None)
    metric_field = next((field for field in fieldnames if normalize_structure_name(field) == "METRIC"), None)
    goal_field = next((field for field in fieldnames if normalize_structure_name(field) == "GOAL"), None)
    value_field = next((field for field in fieldnames if normalize_structure_name(field) == "VALUE"), None)
    fallback_field = fieldnames[0] if fieldnames else None

    def unpack_cell(cell_payload: object) -> Tuple[object, str]:
        if isinstance(cell_payload, dict):
            return cell_payload.get("value"), str(cell_payload.get("number_format", "") or "")
        return cell_payload, ""

    def format_numeric_text(value: float) -> str:
        rounded = round(value)
        if np.isclose(value, rounded):
            return str(int(rounded))
        return f"{value:.6g}"

    def format_value_text(cell_payload: object, metric_text: str) -> str:
        raw_value, number_format = unpack_cell(cell_payload)
        if raw_value is None:
            return ""
        if isinstance(raw_value, str):
            return raw_value.strip()
        if isinstance(raw_value, (int, float, np.integer, np.floating)):
            numeric_value = float(raw_value)
            metric_key = metric_text.strip().upper().replace(" ", "")
            is_percent_value = "%" in number_format or (
                metric_key.startswith("V") and 0.0 <= numeric_value <= 1.0
            )
            if is_percent_value:
                return f"{format_numeric_text(numeric_value * 100.0)}%"
            return format_numeric_text(numeric_value)
        return str(raw_value).strip()

    def split_structure_names(raw_name_text: str) -> List[str]:
        split_names = [part.strip() for part in re.split(r",", raw_name_text) if part.strip()]
        expanded_names: List[str] = []
        for name in split_names:
            match = re.fullmatch(r"(.+)_L/R", name, flags=re.IGNORECASE)
            if match is not None:
                prefix = match.group(1).strip()
                if prefix:
                    expanded_names.extend([f"{prefix}_L", f"{prefix}_R"])
                continue
            expanded_names.append(name)
        if expanded_names:
            return expanded_names
        stripped = raw_name_text.strip()
        return [stripped] if stripped else []

    for row in rows:
        raw_name = ""
        if oar_field:
            raw_name, _ = unpack_cell(row.get(oar_field, ""))
            raw_name = str(raw_name or "")
        elif fallback_field:
            raw_name, _ = unpack_cell(row.get(fallback_field, ""))
            raw_name = str(raw_name or "")
        structure_names = split_structure_names(raw_name)
        if structure_names:
            metric_value, _ = unpack_cell(row.get(metric_field, "")) if metric_field else ("", "")
            comparator_value, _ = unpack_cell(row.get(goal_field, "")) if goal_field else ("", "")
            metric = str(metric_value or "").strip()
            comparator = str(comparator_value or "").strip()
            value_text = format_value_text(row.get(value_field, ""), metric) if value_field else ""
            for structure_name in structure_names:
                normalized = normalize_structure_name(structure_name)
                if not normalized:
                    continue
                if normalized not in allowed_names:
                    structure_order.append(normalized)
                allowed_names.add(normalized)
                if metric or comparator or value_text:
                    goals_by_structure.setdefault(normalized, []).append(
                        StructureGoal(
                            structure_name=structure_name,
                            metric=metric,
                            comparator=comparator,
                            value_text=value_text,
                        )
                    )

    return allowed_names, goals_by_structure, structure_order


def list_constraints_workbook_sheets(path: str) -> List[str]:
    workbook_path = Path(path)
    stat_result = workbook_path.stat()
    cache_key = (str(workbook_path.resolve()), int(stat_result.st_mtime_ns), int(stat_result.st_size))
    cached_sheet_names = _CONSTRAINT_WORKBOOK_SHEETS_CACHE.get(cache_key)
    if cached_sheet_names is not None:
        return list(cached_sheet_names)

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    _CONSTRAINT_WORKBOOK_SHEETS_CACHE.clear()
    _CONSTRAINT_WORKBOOK_SHEETS_CACHE[cache_key] = list(sheet_names)
    return sheet_names


def _cell_payload(cell) -> Dict[str, object]:
    return {
        "value": cell.value,
        "number_format": cell.number_format,
    }


def _cell_text(cell_payload: object) -> str:
    if isinstance(cell_payload, dict):
        value = cell_payload.get("value")
    else:
        value = cell_payload
    if value is None:
        return ""
    return str(value).strip()


def _normalize_constraint_header(cell_payload: object) -> str:
    return normalize_structure_name(_cell_text(cell_payload))


def _extract_plan_dose_per_fraction_values(plan_phases: Optional[List[RTPlanPhase]]) -> List[float]:
    values: List[float] = []
    if not plan_phases:
        return values
    for phase in plan_phases:
        if phase.prescription_dose_gy <= 0.0 or phase.fractions_planned <= 0:
            continue
        dose_per_fraction_gy = phase.prescription_dose_gy / float(phase.fractions_planned)
        if any(abs(existing - dose_per_fraction_gy) <= 0.01 for existing in values):
            continue
        values.append(dose_per_fraction_gy)
    return values


def _extract_plan_fraction_counts(plan_phases: Optional[List[RTPlanPhase]]) -> List[int]:
    values: List[int] = []
    if not plan_phases:
        return values
    for phase in plan_phases:
        fractions_planned = int(phase.fractions_planned)
        if fractions_planned <= 0:
            continue
        if fractions_planned in values:
            continue
        values.append(fractions_planned)
    return values


def _parse_constraint_block_label(label_text: str) -> Tuple[str, Optional[float]]:
    stripped = label_text.strip()
    if not stripped:
        return "", None
    if normalize_structure_name(stripped) in {"NA", "N/A"}:
        return "na", None
    match = re.search(r"\bD\s*/\s*F\b\s*([0-9]+(?:\.[0-9]+)?)", stripped, flags=re.IGNORECASE)
    if match is not None:
        try:
            return "dose_per_fraction", float(match.group(1))
        except ValueError:
            return "", None
    match = re.search(r"\bF\b\s*([0-9]+(?:\.[0-9]+)?)", stripped, flags=re.IGNORECASE)
    if match is not None:
        try:
            return "fraction_count", float(match.group(1))
        except ValueError:
            return "", None
    return "", None


def _extract_constraints_table_blocks(worksheet) -> List[Dict[str, object]]:
    rows = list(worksheet.iter_rows(values_only=False))
    if len(rows) < 2:
        return []

    max_cols = max(len(row) for row in rows)
    label_row_payloads = [_cell_payload(cell) for cell in rows[0]]
    header_row_payloads = [_cell_payload(cell) for cell in rows[1]]

    blocks: List[Dict[str, object]] = []
    col = 0
    while col + 3 < max_cols:
        headers = [
            _normalize_constraint_header(header_row_payloads[col + offset] if col + offset < len(header_row_payloads) else None)
            for offset in range(4)
        ]
        if headers == ["OAR", "METRIC", "GOAL", "VALUE"]:
            fieldnames = [
                _cell_text(header_row_payloads[col + offset] if col + offset < len(header_row_payloads) else None)
                for offset in range(4)
            ]
            block_rows: List[Dict[str, object]] = []
            for row in rows[2:]:
                payloads = [_cell_payload(cell) for cell in row]
                row_payload: Dict[str, object] = {}
                has_any_value = False
                for offset, fieldname in enumerate(fieldnames):
                    payload = payloads[col + offset] if col + offset < len(payloads) else None
                    row_payload[fieldname] = payload
                    if _cell_text(payload):
                        has_any_value = True
                if has_any_value:
                    block_rows.append(row_payload)
            label_payload = label_row_payloads[col] if col < len(label_row_payloads) else None
            blocks.append(
                {
                    "label": _cell_text(label_payload),
                    "fieldnames": fieldnames,
                    "rows": block_rows,
                }
            )
            col += 4
            continue
        col += 1

    return blocks


def _select_constraints_table_block(
    blocks: List[Dict[str, object]],
    plan_phases: Optional[List[RTPlanPhase]],
) -> Optional[Dict[str, object]]:
    if not blocks:
        return None
    if len(blocks) == 1:
        return blocks[0]

    dose_per_fraction_values = _extract_plan_dose_per_fraction_values(plan_phases)
    fraction_counts = _extract_plan_fraction_counts(plan_phases)
    fallback_block: Optional[Dict[str, object]] = None
    matching_blocks: List[Tuple[float, Dict[str, object]]] = []

    for block in blocks:
        label_kind, label_value = _parse_constraint_block_label(str(block.get("label", "")))
        if label_kind == "na":
            fallback_block = block
        elif label_kind == "dose_per_fraction" and label_value is not None:
            for dose_per_fraction_gy in dose_per_fraction_values:
                if abs(label_value - dose_per_fraction_gy) <= 0.05:
                    matching_blocks.append((abs(label_value - dose_per_fraction_gy), block))
                    break
        elif label_kind == "fraction_count" and label_value is not None:
            for fraction_count in fraction_counts:
                if abs(label_value - float(fraction_count)) <= 0.05:
                    matching_blocks.append((abs(label_value - float(fraction_count)), block))
                    break

    if matching_blocks:
        matching_blocks.sort(key=lambda item: item[0])
        return matching_blocks[0][1]
    if fallback_block is not None:
        return fallback_block
    return blocks[0]


def load_structure_constraints_sheet(
    path: str,
    sheet_name: str,
    plan_phases: Optional[List[RTPlanPhase]] = None,
) -> Tuple[set[str], dict[str, List[StructureGoal]], List[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Constraints sheet '{sheet_name}' was not found in {Path(path).name}.")
        worksheet = workbook[sheet_name]
        blocks = _extract_constraints_table_blocks(worksheet)
        selected_block = _select_constraints_table_block(blocks, plan_phases)
        if selected_block is not None:
            return _parse_structure_goal_rows(
                list(selected_block.get("fieldnames", [])),
                list(selected_block.get("rows", [])),
            )

        row_iter = worksheet.iter_rows(values_only=False)
        header_row = next(row_iter, None)
        if header_row is None:
            return set(), {}, []
        fieldnames = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
        rows: List[Dict[str, object]] = []
        for cells in row_iter:
            row: Dict[str, object] = {}
            for index, fieldname in enumerate(fieldnames):
                if index < len(cells):
                    row[fieldname] = _cell_payload(cells[index])
                else:
                    row[fieldname] = None
            rows.append(row)
        return _parse_structure_goal_rows(fieldnames, rows)
    finally:
        workbook.close()

def find_constraint_script_xml_file(folder: Optional[str]) -> Optional[str]:
    if not folder:
        return None
    try:
        entries = [path for path in Path(folder).iterdir() if path.is_file()]
    except OSError:
        return None
    if not entries:
        return None

    json_entries = [
        path
        for path in entries
        if path.suffix.lower() == ".json"
        and path.name.lower() in {"constraints.json", "contraints.json"}
    ]
    if json_entries:
        json_entries.sort(key=lambda candidate: (candidate.name.lower() != "contraints.json", candidate.name.lower()))
        return str(json_entries[0])

    fuzzy_json_entries = [
        path
        for path in entries
        if path.suffix.lower() == ".json"
        and ("constraint" in path.name.lower() or "contraint" in path.name.lower())
    ]
    if fuzzy_json_entries:
        fuzzy_json_entries.sort(key=lambda candidate: candidate.name.lower())
        return str(fuzzy_json_entries[0])

    entries = [path for path in entries if path.suffix.lower() == ".xml"]
    if not entries:
        return None
    entries.sort(
        key=lambda candidate: (
            0 if candidate.name.lower().endswith("_ctable.xml") else 1,
            candidate.name.lower(),
        )
    )
    return str(entries[0])


def _format_script_goal_number(value_text: str) -> str:
    try:
        numeric_value = float(value_text)
    except (TypeError, ValueError):
        return str(value_text).strip()
    rounded_value = round(numeric_value)
    if abs(numeric_value - rounded_value) <= 1e-9:
        return str(int(rounded_value))
    return f"{numeric_value:.6g}"


def _normalize_script_goal_value(goal_text: str) -> str:
    text = " ".join(str(goal_text or "").split())
    if not text:
        return ""
    match = re.search(r"([-+]?\d*\.?\d+)\s*(%|CC|CM3|GY)", text, flags=re.IGNORECASE)
    if match is not None:
        numeric_text = _format_script_goal_number(match.group(1))
        unit = match.group(2).upper()
        if unit == "%":
            return f"{numeric_text}%"
        if unit == "CM3":
            unit = "CC"
        if unit == "CC":
            return f"{numeric_text} cc"
        if unit == "GY":
            return f"{numeric_text} Gy"
        return f"{numeric_text} {unit}"
    match = re.search(r"[-+]?\d*\.?\d+", text)
    if match is None:
        return text
    return _format_script_goal_number(match.group(0))


def _parse_script_constraint_clause(constraint_text: str) -> Optional[Tuple[str, str]]:
    text = " ".join(str(constraint_text or "").replace("≤", "<=").replace("≥", ">=").split())
    if not text:
        return None
    match = re.search(r"(<=|>=|==|=|<|>)\s*$", text)
    if match is None:
        return None
    metric = text[:match.start()].strip().replace(" ", "")
    comparator = match.group(1).strip()
    if not metric or not comparator:
        return None
    return metric, comparator


def _build_script_constraint_note_text(result_text: str, comment_text: str) -> str:
    pieces: List[str] = []
    cleaned_result = str(result_text or "").strip()
    cleaned_comment = str(comment_text or "").strip()
    if cleaned_result:
        pieces.append(f"Eclipse: {cleaned_result}")
    if cleaned_comment:
        pieces.append(f"Comment: {cleaned_comment}")
    return "    ".join(pieces)


def _build_script_constraint_note_key(
    normalized_name: str,
    metric: str,
    comparator: str,
    value_text: str,
) -> str:
    return "||".join(
        [
            normalized_name,
            metric.strip(),
            comparator.strip(),
            value_text.strip(),
        ]
    )


def load_structure_constraints_script(
    path: str,
) -> Tuple[set[str], dict[str, List[StructureGoal]], List[str], dict[str, str]]:
    if Path(path).suffix.lower() == ".json":
        return load_structure_constraints_json(path)

    try:
        root = ET.parse(path).getroot()
    except (ET.ParseError, OSError) as exc:
        raise ValueError(f"Failed to read script constraints from {Path(path).name}: {exc}") from exc

    allowed_names: set[str] = set()
    goals_by_structure: dict[str, List[StructureGoal]] = {}
    structure_order: List[str] = []
    note_text_by_goal_key: dict[str, str] = {}
    seen_goals: set[Tuple[str, str, str, str]] = set()

    for item in root.findall('.//Constraints_x0020_Checks_x0020_2'):
        display_name = str(item.findtext('strTemp', '') or '').strip()
        plan_name = str(item.findtext('strPlan', '') or '').strip()
        normalized_display_name = normalize_structure_name(display_name)
        normalized_plan_name = normalize_structure_name(plan_name)
        if normalized_display_name.startswith('PTV') or normalized_plan_name.startswith('PTV'):
            continue

        normalized_name = normalized_plan_name or normalized_display_name
        structure_name = plan_name or display_name
        if not normalized_name or not structure_name:
            continue

        if normalized_name not in allowed_names:
            structure_order.append(normalized_name)
        allowed_names.add(normalized_name)

        note_text = _build_script_constraint_note_text(
            item.findtext('cPlan', ''),
            item.findtext('cComment', ''),
        )
        for constraint_field, goal_field in (('constraint', 'cGoal'), ('constraint2', 'cGoal2')):
            parsed_constraint = _parse_script_constraint_clause(item.findtext(constraint_field, ''))
            value_text = _normalize_script_goal_value(item.findtext(goal_field, ''))
            if parsed_constraint is None or not value_text:
                continue
            metric, comparator = parsed_constraint
            goal_key = (normalized_name, metric.upper(), comparator, value_text.upper())
            if goal_key in seen_goals:
                continue
            seen_goals.add(goal_key)
            goal = StructureGoal(
                structure_name=structure_name,
                metric=metric,
                comparator=comparator,
                value_text=value_text,
            )
            goals_by_structure.setdefault(normalized_name, []).append(goal)
            if note_text:
                note_text_by_goal_key[_build_script_constraint_note_key(
                    normalized_name,
                    goal.metric,
                    goal.comparator,
                    goal.value_text,
                )] = note_text

    return allowed_names, goals_by_structure, structure_order, note_text_by_goal_key


def load_structure_constraints_json(
    path: str,
) -> Tuple[set[str], dict[str, List[StructureGoal]], List[str], dict[str, str]]:
    try:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Failed to read script JSON constraints from {Path(path).name}: {exc}") from exc

    if not isinstance(payload, list):
        raise ValueError(f"Script JSON constraints in {Path(path).name} must contain a top-level list.")

    allowed_names: set[str] = set()
    goals_by_structure: dict[str, List[StructureGoal]] = {}
    structure_order: List[str] = []
    note_text_by_goal_key: dict[str, str] = {}
    seen_goals: set[Tuple[str, str, str, str]] = set()

    for item in payload:
        if not isinstance(item, dict):
            continue

        display_name = str(item.get("strTemp", "") or "").strip()
        plan_name = str(item.get("strPlan", "") or "").strip()
        normalized_display_name = normalize_structure_name(display_name)
        normalized_plan_name = normalize_structure_name(plan_name)
        if normalized_display_name.startswith("PTV") or normalized_plan_name.startswith("PTV"):
            continue

        normalized_name = normalized_plan_name or normalized_display_name
        structure_name = plan_name or display_name
        if not normalized_name or not structure_name:
            continue

        if normalized_name not in allowed_names:
            structure_order.append(normalized_name)
        allowed_names.add(normalized_name)

        note_text = _build_script_constraint_note_text(
            item.get("cPlan", ""),
            item.get("cComment", ""),
        )
        for constraint_field, goal_field in (("constraint", "cGoal"), ("constraint2", "cGoal2")):
            parsed_constraint = _parse_script_constraint_clause(str(item.get(constraint_field, "") or ""))
            value_text = _normalize_script_goal_value(str(item.get(goal_field, "") or ""))
            if parsed_constraint is None or not value_text:
                continue
            metric, comparator = parsed_constraint
            goal_key = (normalized_name, metric.upper(), comparator, value_text.upper())
            if goal_key in seen_goals:
                continue
            seen_goals.add(goal_key)
            goal = StructureGoal(
                structure_name=structure_name,
                metric=metric,
                comparator=comparator,
                value_text=value_text,
            )
            goals_by_structure.setdefault(normalized_name, []).append(goal)
            if note_text:
                note_text_by_goal_key[_build_script_constraint_note_key(
                    normalized_name,
                    goal.metric,
                    goal.comparator,
                    goal.value_text,
                )] = note_text

    return allowed_names, goals_by_structure, structure_order, note_text_by_goal_key


def _summarize_rtplan_phase_records(
    phase_records: List[Dict[str, object]],
) -> Optional[Tuple[str, ...]]:
    if not phase_records:
        return None

    patient_name = ""
    patient_id = ""
    total_prescription_dose_gy = 0.0
    total_fractions = 0

    for record in phase_records:
        if not patient_name:
            patient_name = str(record.get("patient_name", "")).strip()
        if not patient_id:
            patient_id = str(record.get("patient_id", "")).strip()
        prescription_dose_gy = float(record.get("prescription_dose_gy", 0.0) or 0.0)
        fractions_planned = int(record.get("fractions_planned", 0) or 0)
        if prescription_dose_gy > 0.0:
            total_prescription_dose_gy += prescription_dose_gy
        if fractions_planned > 0:
            total_fractions += fractions_planned

    if not patient_name and not patient_id and total_prescription_dose_gy <= 0.0 and total_fractions <= 0:
        return None

    line_1 = patient_name or "Patient name unavailable"
    line_2 = f"ID: {patient_id}" if patient_id else "ID unavailable"
    if total_prescription_dose_gy > 0.0 and total_fractions > 0:
        dose_per_fraction_gy = total_prescription_dose_gy / float(total_fractions)
        line_3 = f"{total_prescription_dose_gy:.2f} Gy | {total_fractions} fx | {dose_per_fraction_gy:.2f} Gy/fx"
    elif total_prescription_dose_gy > 0.0:
        line_3 = f"{total_prescription_dose_gy:.2f} Gy"
    elif total_fractions > 0:
        line_3 = f"{total_fractions} fx"
    else:
        line_3 = "Prescription unavailable"

    plan_count = len(phase_records)
    if plan_count > 1:
        line_4 = f"{plan_count} phases"
        return (line_1, line_2, line_3, line_4)
    return (line_1, line_2, line_3)


def scan_patient_folder(
    folder: str,
) -> PatientFileDiscovery:
    ct_series_by_uid: Dict[str, Dict[str, object]] = {}
    rtstruct_records: List[Dict[str, object]] = []
    rtdose_records: List[Dict[str, object]] = []
    rtplan_phase_records: List[Dict[str, object]] = []
    registration_paths: List[str] = []
    registration_transforms: Dict[Tuple[str, str], np.ndarray] = {}
    dose_record_by_plan_uid: Dict[str, Dict[str, object]] = {}

    for path in sorted(Path(folder).rglob("*")):
        if not path.is_file():
            continue

        try:
            ds = pydicom.dcmread(str(path), stop_before_pixels=True, force=True)
            modality = str(safe_get(ds, "Modality", "")).upper()
            dataset_datetime = _get_dataset_datetime(ds)
        except Exception as exc:
            logger.warning("Skipping unreadable DICOM during folder scan: %s (%s)", path, exc)
            continue

        if modality == "CT":
            series_uid = str(safe_get(ds, "SeriesInstanceUID", "")).strip() or str(path)
            ct_record = ct_series_by_uid.setdefault(
                series_uid,
                {
                    "series_uid": series_uid,
                    "paths": [],
                    "frame_of_reference_uid": str(safe_get(ds, "FrameOfReferenceUID", "")).strip(),
                    "study_uid": str(safe_get(ds, "StudyInstanceUID", "")).strip(),
                    "datetime": dataset_datetime,
                },
            )
            cast_paths = ct_record.setdefault("paths", [])
            if isinstance(cast_paths, list):
                cast_paths.append(str(path))
            if ct_record.get("datetime") is None and dataset_datetime is not None:
                ct_record["datetime"] = dataset_datetime
        elif modality == "RTSTRUCT":
            rtstruct_records.append(
                {
                    "path": str(path),
                    "sop_instance_uid": str(safe_get(ds, "SOPInstanceUID", "")).strip(),
                    "frame_of_reference_uid": str(safe_get(ds, "FrameOfReferenceUID", "")).strip(),
                    "referenced_ct_series_uid": _extract_referenced_ct_series_uid(ds),
                    "datetime": dataset_datetime,
                }
            )
        elif modality == "RTDOSE":
            record = {
                "path": str(path),
                "frame_of_reference_uid": str(safe_get(ds, "FrameOfReferenceUID", "")).strip(),
                "datetime": dataset_datetime,
                "referenced_rtplan_uid": "",
            }
            for item in safe_get(ds, "ReferencedRTPlanSequence", []):
                referenced_uid = str(safe_get(item, "ReferencedSOPInstanceUID", "")).strip()
                if referenced_uid:
                    record["referenced_rtplan_uid"] = referenced_uid
                    existing_record = dose_record_by_plan_uid.get(referenced_uid)
                    if existing_record is None or _datetime_sort_key(existing_record.get("datetime")) < _datetime_sort_key(dataset_datetime):
                        dose_record_by_plan_uid[referenced_uid] = record
                    break
            rtdose_records.append(record)
        elif modality == "RTPLAN":
            sop_instance_uid = str(safe_get(ds, "SOPInstanceUID", "")).strip()
            prescription_doses_gy = _extract_rtplan_prescription_doses_gy(ds)
            rtplan_phase_records.append(
                {
                    "path": str(path),
                    "sop_instance_uid": sop_instance_uid,
                    "prescription_dose_gy": max(prescription_doses_gy) if prescription_doses_gy else 0.0,
                    "fractions_planned": _extract_rtplan_number_of_fractions(ds),
                    "target_structure_name": _extract_rtplan_target_structure_name(ds),
                    "plan_label": str(safe_get(ds, "RTPlanLabel", "")).strip(),
                    "plan_name": str(safe_get(ds, "RTPlanName", "")).strip(),
                    "patient_name": _format_patient_name(safe_get(ds, "PatientName", "")),
                    "patient_id": str(safe_get(ds, "PatientID", "")).strip(),
                    "referenced_rtstruct_uid": _extract_referenced_rtstruct_uid(ds),
                    "datetime": dataset_datetime,
                }
            )
        elif modality == "REG":
            registration_paths.append(str(path))
            for transform_key, transform in _load_registration_transforms(str(path)).items():
                registration_transforms[transform_key] = transform

    sorted_ct_records = sorted(
        ct_series_by_uid.values(),
        key=lambda record: (_datetime_sort_key(record.get("datetime")), str(record.get("series_uid", ""))),
    )
    sorted_rtstruct_records = sorted(
        rtstruct_records,
        key=lambda record: (_datetime_sort_key(record.get("datetime")), str(record.get("path", ""))),
    )
    sorted_rtplan_records = sorted(
        rtplan_phase_records,
        key=lambda record: (_datetime_sort_key(record.get("datetime")), str(record.get("path", ""))),
    )
    sorted_rtdose_records = sorted(
        rtdose_records,
        key=lambda record: (_datetime_sort_key(record.get("datetime")), str(record.get("path", ""))),
    )

    primary_rtstruct_record: Optional[Dict[str, object]] = None
    primary_ct_record: Optional[Dict[str, object]] = None
    for rtstruct_record in reversed(sorted_rtstruct_records):
        referenced_ct_series_uid = str(rtstruct_record.get("referenced_ct_series_uid", "")).strip()
        if referenced_ct_series_uid and referenced_ct_series_uid in ct_series_by_uid:
            primary_rtstruct_record = rtstruct_record
            primary_ct_record = ct_series_by_uid[referenced_ct_series_uid]
            break
        frame_of_reference_uid = str(rtstruct_record.get("frame_of_reference_uid", "")).strip()
        if frame_of_reference_uid:
            matching_ct_records = [
                record
                for record in sorted_ct_records
                if str(record.get("frame_of_reference_uid", "")).strip() == frame_of_reference_uid
            ]
            if matching_ct_records:
                primary_rtstruct_record = rtstruct_record
                primary_ct_record = matching_ct_records[-1]
                break

    if primary_ct_record is None and sorted_ct_records:
        primary_ct_record = sorted_ct_records[-1]

    if primary_rtstruct_record is None and primary_ct_record is not None:
        primary_frame_of_reference_uid = str(primary_ct_record.get("frame_of_reference_uid", "")).strip()
        matching_rtstruct_records = [
            record
            for record in sorted_rtstruct_records
            if str(record.get("frame_of_reference_uid", "")).strip() == primary_frame_of_reference_uid
        ]
        if matching_rtstruct_records:
            primary_rtstruct_record = matching_rtstruct_records[-1]

    primary_ct_paths = list(primary_ct_record.get("paths", [])) if primary_ct_record is not None else []
    primary_rtstruct_path = str(primary_rtstruct_record.get("path", "")).strip() or None
    primary_ct_frame_of_reference_uid = (
        str(primary_ct_record.get("frame_of_reference_uid", "")).strip() if primary_ct_record is not None else ""
    )

    plan_phases = [
        RTPlanPhase(
            sop_instance_uid=str(record.get("sop_instance_uid", "")),
            prescription_dose_gy=float(record.get("prescription_dose_gy", 0.0) or 0.0),
            fractions_planned=int(record.get("fractions_planned", 0) or 0),
            dose_path=str(
                (dose_record_by_plan_uid.get(str(record.get("sop_instance_uid", "")), {}) or {}).get("path", "")
            ),
            target_structure_name=str(record.get("target_structure_name", "")),
            plan_label=str(record.get("plan_label", "")),
            plan_name=str(record.get("plan_name", "")),
        )
        for record in sorted_rtplan_records
    ]
    rtdose_paths = [str(record.get("path", "")) for record in sorted_rtdose_records if str(record.get("path", ""))]
    rtplan_paths = [str(record.get("path", "")) for record in sorted_rtplan_records if str(record.get("path", ""))]
    dose_path_to_ct_transform_by_path: Dict[str, np.ndarray] = {}
    if primary_ct_frame_of_reference_uid:
        for dose_record in sorted_rtdose_records:
            dose_path = str(dose_record.get("path", "")).strip()
            dose_frame_of_reference_uid = str(dose_record.get("frame_of_reference_uid", "")).strip()
            if not dose_path or not dose_frame_of_reference_uid or dose_frame_of_reference_uid == primary_ct_frame_of_reference_uid:
                continue
            transform = registration_transforms.get((primary_ct_frame_of_reference_uid, dose_frame_of_reference_uid))
            if transform is None:
                raise ValueError(
                    "RTDOSE frame of reference does not match the primary CT and no REG transform was found "
                    f"for dose '{Path(dose_path).name}'."
                )
            dose_path_to_ct_transform_by_path[dose_path] = np.asarray(transform, dtype=np.float64)

    return PatientFileDiscovery(
        ct_paths=primary_ct_paths,
        rtstruct_path=primary_rtstruct_path,
        rtdose_paths=rtdose_paths,
        rtplan_paths=rtplan_paths,
        registration_paths=sorted(registration_paths),
        dose_path_to_ct_transform_by_path=dose_path_to_ct_transform_by_path,
        plan_phases=plan_phases,
        patient_plan_lines=_summarize_rtplan_phase_records(sorted_rtplan_records),
    )

def load_ct_series_from_paths(ct_paths: List[str]) -> CTVolume:
    files = []
    for path in ct_paths:
        try:
            ds = pydicom.dcmread(path, stop_before_pixels=False, force=True)
            if safe_get(ds, "Modality", "") == "CT":
                files.append(ds)
        except Exception as exc:
            logger.warning("Skipping CT candidate during CT load: %s (%s)", path, exc)

    if not files:
        raise ValueError("No CT DICOM slices found in the selected folder.")

    first_iop = get_iop(files[0])
    row_cos, col_cos, normal = get_ct_row_col_normal(first_iop)

    def slice_sort_key(ds):
        ipp = get_ipp(ds)
        return float(np.dot(ipp, normal))

    files.sort(key=slice_sort_key)

    first = files[0]
    rows = int(first.Rows)
    cols = int(first.Columns)

    px_spacing = np.array([float(x) for x in first.PixelSpacing], dtype=float)
    sy = float(px_spacing[0])
    sx = float(px_spacing[1])

    slice_origins = []
    slice_positions_along_normal = []
    slices = []

    for ds in files:
        arr = ds.pixel_array.astype(np.float32)
        arr = apply_rescale(arr, ds).astype(np.float32)
        slices.append(arr)

        ipp = get_ipp(ds)
        slice_origins.append(ipp)
        slice_positions_along_normal.append(float(np.dot(ipp, normal)))

    volume = np.stack(slices, axis=0)
    slice_origins = np.asarray(slice_origins, dtype=float)
    slice_positions_along_normal = np.asarray(slice_positions_along_normal, dtype=float)

    if len(slice_positions_along_normal) > 1:
        dz = float(np.median(np.diff(slice_positions_along_normal)))
    else:
        dz = float(safe_get(first, "SliceThickness", 1.0))

    return CTVolume(
        volume_hu=volume,
        slice_origins_xyz_mm=slice_origins,
        z_positions_mm=slice_positions_along_normal,
        spacing_xyz_mm=np.array([sx, sy, abs(dz)], dtype=float),
        image_orientation_patient=first_iop,
        study_uid=str(safe_get(first, "StudyInstanceUID", "")),
        frame_of_reference_uid=str(safe_get(first, "FrameOfReferenceUID", "")),
        rows=rows,
        cols=cols,
    )

def _format_patient_name(name_value: object) -> str:
    text = str(name_value or "").strip()
    if not text:
        return ""
    parts = [part.strip() for part in text.split("^")]
    if len(parts) > 1:
        family = parts[0] if len(parts) > 0 else ""
        given = parts[1] if len(parts) > 1 else ""
        middle = parts[2] if len(parts) > 2 else ""
        prefix = parts[3] if len(parts) > 3 else ""
        suffix = parts[4] if len(parts) > 4 else ""
        reordered = " ".join(
            part
            for part in [prefix, given, middle, family, suffix]
            if part
        )
        if reordered:
            return reordered
    return " ".join(part for part in text.replace("^", " ").split() if part)


def _extract_rtplan_prescription_doses_gy(ds: pydicom.dataset.Dataset) -> List[float]:
    prescription_doses: List[float] = []
    for item in safe_get(ds, "DoseReferenceSequence", []):
        value = safe_get(item, "TargetPrescriptionDose", None)
        if value in {None, ""}:
            continue
        try:
            prescription_doses.append(float(value))
        except (TypeError, ValueError):
            continue

    return prescription_doses


def _extract_rtplan_number_of_fractions(ds: pydicom.dataset.Dataset) -> int:
    total_fractions = 0
    for item in safe_get(ds, "FractionGroupSequence", []):
        value = safe_get(item, "NumberOfFractionsPlanned", None)
        if value in {None, ""}:
            continue
        try:
            total_fractions += int(value)
        except (TypeError, ValueError):
            continue

    return total_fractions


def _extract_rtplan_target_structure_name(ds: pydicom.dataset.Dataset) -> str:
    for item in safe_get(ds, "DoseReferenceSequence", []):
        description = str(safe_get(item, "DoseReferenceDescription", "")).strip()
        if not description:
            continue
        normalized_description = normalize_structure_name(description)
        ptv_index = normalized_description.find("PTV")
        if ptv_index >= 0:
            return normalized_description[ptv_index:]
    return ""

def load_rtdose(path: str) -> DoseVolume:
    ds = pydicom.dcmread(path, stop_before_pixels=False)
    if safe_get(ds, "Modality", "") != "RTDOSE":
        raise ValueError("Selected file is not an RTDOSE object.")

    arr = ds.pixel_array.astype(np.float32)
    dose_grid_scaling = float(safe_get(ds, "DoseGridScaling", 1.0))
    arr = arr * dose_grid_scaling

    if arr.ndim == 2:
        arr = arr[np.newaxis, :, :]

    px_spacing = np.array([float(x) for x in ds.PixelSpacing], dtype=float)
    sy = float(px_spacing[0])
    sx = float(px_spacing[1])

    iop = get_iop(ds)
    _, _, dose_normal = get_ct_row_col_normal(iop)
    ipp = get_ipp(ds)
    offsets = np.array(safe_get(ds, "GridFrameOffsetVector", list(range(arr.shape[0]))), dtype=float)
    slice_origins = ipp[None, :] + offsets[:, None] * dose_normal[None, :]
    z_positions = slice_origins @ dose_normal

    dz = float(np.median(np.diff(z_positions))) if len(z_positions) > 1 else 1.0

    return DoseVolume(
        dose_gy=arr,
        slice_origins_xyz_mm=slice_origins,
        z_positions_mm=z_positions,
        origin_xyz_mm=slice_origins[0].copy(),
        spacing_xyz_mm=np.array([sx, sy, abs(dz)], dtype=float),
        image_orientation_patient=iop,
        frame_of_reference_uid=str(safe_get(ds, "FrameOfReferenceUID", "")),
        dose_units=str(safe_get(ds, "DoseUnits", "")),
    )


def build_ct_aligned_dose_volume(
    ct: CTVolume,
    dose_gy: np.ndarray,
    *,
    dose_units: str = "",
) -> DoseVolume:
    dose_array = np.asarray(dose_gy, dtype=np.float32)
    if dose_array.shape != ct.volume_hu.shape:
        raise ValueError("CT-aligned dose array shape does not match the CT volume geometry.")
    return DoseVolume(
        dose_gy=dose_array.copy(),
        slice_origins_xyz_mm=np.asarray(ct.slice_origins_xyz_mm, dtype=np.float32).copy(),
        z_positions_mm=np.asarray(ct.z_positions_mm, dtype=np.float32).copy(),
        origin_xyz_mm=np.asarray(ct.slice_origins_xyz_mm[0], dtype=np.float32).copy(),
        spacing_xyz_mm=np.asarray(ct.spacing_xyz_mm, dtype=np.float32).copy(),
        image_orientation_patient=np.asarray(ct.image_orientation_patient, dtype=np.float32).copy(),
        frame_of_reference_uid=ct.frame_of_reference_uid,
        dose_units=str(dose_units or ""),
    )


def _grid_axis_step_vectors(
    slice_origins_xyz_mm: np.ndarray,
    image_orientation_patient: np.ndarray,
    spacing_xyz_mm: np.ndarray,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    row_cos, col_cos, normal = get_ct_row_col_normal(image_orientation_patient)
    if slice_origins_xyz_mm.shape[0] > 1:
        slice_step_vector = np.median(np.diff(slice_origins_xyz_mm, axis=0), axis=0).astype(np.float64, copy=False)
    else:
        slice_step_vector = normal.astype(np.float64, copy=False) * float(spacing_xyz_mm[2])
    row_step_vector = col_cos.astype(np.float64, copy=False) * float(spacing_xyz_mm[1])
    col_step_vector = row_cos.astype(np.float64, copy=False) * float(spacing_xyz_mm[0])
    return slice_step_vector, row_step_vector, col_step_vector


def resample_dose_to_ct_grid(
    ct: CTVolume,
    dose: DoseVolume,
    *,
    ct_to_dose_transform: Optional[np.ndarray] = None,
) -> np.ndarray:
    if (
        ct_to_dose_transform is None
        and dose.dose_gy.shape == ct.volume_hu.shape
        and dose.slice_origins_xyz_mm.shape == ct.slice_origins_xyz_mm.shape
        and np.allclose(dose.slice_origins_xyz_mm, ct.slice_origins_xyz_mm, atol=1e-3)
        and np.allclose(dose.spacing_xyz_mm, ct.spacing_xyz_mm, atol=1e-6)
        and np.allclose(dose.image_orientation_patient, ct.image_orientation_patient, atol=1e-6)
    ):
        return np.asarray(dose.dose_gy, dtype=np.float32).copy()

    if scipy_affine_transform is not None:
        dose_sort_indices = np.argsort(np.asarray(dose.z_positions_mm, dtype=np.float64))
        dose_volume = np.asarray(dose.dose_gy, dtype=np.float32)[dose_sort_indices]
        dose_slice_origins = np.asarray(dose.slice_origins_xyz_mm, dtype=np.float64)[dose_sort_indices]
        ct_slice_origins = np.asarray(ct.slice_origins_xyz_mm, dtype=np.float64)

        ct_slice_step, ct_row_step, ct_col_step = _grid_axis_step_vectors(
            ct_slice_origins,
            np.asarray(ct.image_orientation_patient, dtype=np.float64),
            np.asarray(ct.spacing_xyz_mm, dtype=np.float64),
        )
        dose_slice_step, dose_row_step, dose_col_step = _grid_axis_step_vectors(
            dose_slice_origins,
            np.asarray(dose.image_orientation_patient, dtype=np.float64),
            np.asarray(dose.spacing_xyz_mm, dtype=np.float64),
        )

        ct_basis = np.column_stack([ct_slice_step, ct_row_step, ct_col_step])
        dose_basis = np.column_stack([dose_slice_step, dose_row_step, dose_col_step])
        dose_basis_inverse = np.linalg.inv(dose_basis)

        transform = np.eye(4, dtype=np.float64) if ct_to_dose_transform is None else np.asarray(
            ct_to_dose_transform,
            dtype=np.float64,
        ).reshape(4, 4)
        rotation = transform[:3, :3]
        translation = transform[:3, 3]

        matrix = dose_basis_inverse @ rotation @ ct_basis
        offset = dose_basis_inverse @ (rotation @ ct_slice_origins[0] + translation - dose_slice_origins[0])
        resampled = scipy_affine_transform(
            dose_volume,
            matrix=matrix,
            offset=offset,
            output_shape=ct.volume_hu.shape,
            order=1,
            mode="constant",
            cval=0.0,
            prefilter=False,
        )
        return np.asarray(resampled, dtype=np.float32)

    return np.stack(
        [
            sample_dose_to_ct_slice(
                ct,
                dose,
                slice_index,
                ct_to_dose_transform=ct_to_dose_transform,
            )
            for slice_index in range(ct.volume_hu.shape[0])
        ],
        axis=0,
    ).astype(np.float32, copy=False)


def validate_dose_geometry(reference: DoseVolume, candidate: DoseVolume, path: str):
    if reference.dose_gy.shape != candidate.dose_gy.shape:
        raise ValueError(
            f"RTDOSE file '{path}' does not match the reference dose grid shape "
            f"{reference.dose_gy.shape} != {candidate.dose_gy.shape}."
        )
    if reference.slice_origins_xyz_mm.shape != candidate.slice_origins_xyz_mm.shape:
        raise ValueError(
            f"RTDOSE file '{path}' does not match the reference dose origins shape "
            f"{reference.slice_origins_xyz_mm.shape} != {candidate.slice_origins_xyz_mm.shape}."
        )

    checks = [
        (np.allclose(reference.slice_origins_xyz_mm, candidate.slice_origins_xyz_mm, atol=1e-3), "dose origins"),
        (np.allclose(reference.spacing_xyz_mm, candidate.spacing_xyz_mm, atol=1e-6), "dose spacing"),
        (
            np.allclose(reference.image_orientation_patient, candidate.image_orientation_patient, atol=1e-6),
            "dose orientation",
        ),
    ]
    for passed, label in checks:
        if not passed:
            raise ValueError(f"RTDOSE file '{path}' does not match the reference {label}.")


def load_combined_rtdose(
    paths: List[str],
    *,
    reference_ct: Optional[CTVolume] = None,
    ct_to_dose_transforms: Optional[Dict[str, np.ndarray]] = None,
) -> DoseVolume:
    if not paths:
        raise ValueError("No RTDOSE files were provided.")

    sorted_paths = sorted(paths)
    loaded = [load_rtdose(path) for path in sorted_paths]

    if reference_ct is not None:
        combined_dose_ct = np.zeros(reference_ct.volume_hu.shape, dtype=np.float32)
        dose_units = ""
        for path, dose in zip(sorted_paths, loaded):
            ct_to_dose_transform = None
            if ct_to_dose_transforms is not None:
                ct_to_dose_transform = ct_to_dose_transforms.get(path)
            if ct_to_dose_transform is None and dose.frame_of_reference_uid not in {"", reference_ct.frame_of_reference_uid}:
                raise ValueError(
                    "RTDOSE frame of reference does not match the reference CT and no REG transform was supplied "
                    f"for dose '{Path(path).name}'."
                )
            combined_dose_ct += resample_dose_to_ct_grid(
                reference_ct,
                dose,
                ct_to_dose_transform=ct_to_dose_transform,
            )
            if not dose_units:
                dose_units = dose.dose_units
        return build_ct_aligned_dose_volume(reference_ct, combined_dose_ct, dose_units=dose_units)

    reference = loaded[0]
    combined_dose = reference.dose_gy.copy()

    for path, dose in zip(sorted_paths[1:], loaded[1:]):
        validate_dose_geometry(reference, dose, path)
        combined_dose += dose.dose_gy

    return DoseVolume(
        dose_gy=combined_dose,
        slice_origins_xyz_mm=reference.slice_origins_xyz_mm.copy(),
        z_positions_mm=reference.z_positions_mm.copy(),
        origin_xyz_mm=reference.origin_xyz_mm.copy(),
        spacing_xyz_mm=reference.spacing_xyz_mm.copy(),
        image_orientation_patient=reference.image_orientation_patient.copy(),
        frame_of_reference_uid=reference.frame_of_reference_uid,
        dose_units=reference.dose_units,
    )


def load_rtstruct(path: str, ct: CTVolume) -> RTStructData:
    ds = pydicom.dcmread(path, stop_before_pixels=False)
    if safe_get(ds, "Modality", "") != "RTSTRUCT":
        raise ValueError("Selected file is not an RTSTRUCT object.")

    # The viewer keeps all RTSTRUCT entries so the axial and DVH tabs can
    # decide independently which structures to show or compute.
    row_cos, col_cos, normal = get_ct_row_col_normal(ct.image_orientation_patient)
    inv_sx = 1.0 / max(float(ct.spacing_xyz_mm[0]), 1e-6)
    inv_sy = 1.0 / max(float(ct.spacing_xyz_mm[1]), 1e-6)
    slice_origins = np.asarray(ct.slice_origins_xyz_mm, dtype=np.float32)
    z_positions = np.asarray(ct.z_positions_mm, dtype=np.float32)

    if z_positions.size > 1:
        z_steps = np.diff(z_positions)
        nominal_dz = float(np.median(z_steps))
        use_direct_slice_lookup = abs(nominal_dz) > 1e-6 and np.allclose(z_steps, nominal_dz, atol=1e-3)
    else:
        nominal_dz = 0.0
        use_direct_slice_lookup = False

    def nearest_slice_index_for_contour(contour_xyz: np.ndarray) -> int:
        contour_pos = float(np.mean(contour_xyz @ normal))
        if use_direct_slice_lookup:
            approx_index = int(round((contour_pos - float(z_positions[0])) / nominal_dz))
            return int(np.clip(approx_index, 0, len(z_positions) - 1))

        idx = int(np.searchsorted(z_positions, contour_pos, side="left"))
        if idx <= 0:
            return 0
        if idx >= len(z_positions):
            return len(z_positions) - 1
        if abs(contour_pos - float(z_positions[idx - 1])) <= abs(float(z_positions[idx]) - contour_pos):
            return idx - 1
        return idx

    roi_name_by_number = {}
    for item in safe_get(ds, "StructureSetROISequence", []):
        roi_name_by_number[int(item.ROINumber)] = str(item.ROIName)

    structures: List[StructureSliceContours] = []

    for roi_contour in safe_get(ds, "ROIContourSequence", []):
        roi_num = int(roi_contour.ReferencedROINumber)
        name = roi_name_by_number.get(roi_num, f"ROI {roi_num}")
        color = tuple(int(c) for c in safe_get(roi_contour, "ROIDisplayColor", [255, 255, 0]))
        by_slice = {}

        for contour in safe_get(roi_contour, "ContourSequence", []):
            data = np.asarray(contour.ContourData, dtype=np.float32).reshape(-1, 3)
            if len(data) < 3:
                continue

            k = nearest_slice_index_for_contour(data)
            rel = data - slice_origins[k][None, :]
            cols = rel @ row_cos * inv_sx
            rows = rel @ col_cos * inv_sy
            rc = np.column_stack([rows, cols]).astype(np.float32, copy=False)
            by_slice.setdefault(k, []).append(rc)

        structures.append(
            StructureSliceContours(
                name=name,
                color_rgb=color,
                points_rc_by_slice=by_slice,
            )
        )

    return RTStructData(
        structures=structures,
        frame_of_reference_uid=str(safe_get(ds, "FrameOfReferenceUID", "")),
    )
